import os

project_root = "/mnt/c/Users/khand/GovTrackAI"

files = {
    "excel/generator.py": """import os
import shutil
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference, PieChart
from db.connection import SessionLocal
from db.models import Job, Organization

class ExcelGenerator:
    def __init__(self, output_path="GovTrackAI_Tracker.xlsx", backup_dir="backups"):
        self.output_path = output_path
        self.backup_dir = backup_dir
        os.makedirs(self.backup_dir, exist_ok=True)

    def generate_dashboard(self):
        db = SessionLocal()
        try:
            wb = Workbook()
            ws_data = wb.active
            ws_data.title = "Recruitment Tracker"

            headers = ["ID", "Organization", "Title", "Salary", "Vacancies", "Deadline", "Status", "Applied", "Priority", "Domain", "URL"]
            ws_data.append(headers)

            jobs = db.query(Job, Organization.name.label("org_name")).join(Organization, Job.org_id == Organization.id).filter(Job.is_trashed == False).all()
            
            # Write data
            for job, org in jobs:
                ws_data.append([
                    job.id,
                    org,
                    job.title,
                    job.salary,
                    job.vacancies,
                    job.deadline.strftime("%Y-%m-%d") if job.deadline else "",
                    job.status,
                    "Yes" if job.is_applied else "No",
                    job.priority,
                    job.domain,
                    job.url
                ])

            # Formatting
            for cell in ws_data[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")

            # Table for filters
            if len(jobs) > 0:
                tab = Table(displayName="JobsTable", ref=f"A1:K{len(jobs)+1}")
                style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab.tableStyleInfo = style
                ws_data.add_table(tab)
            
            # Conditional Formatting
            # Red text if deadline is past (assume F is deadline)
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            red_font = Font(color='9C0006')
            ws_data.conditional_formatting.add(
                f'F2:F{len(jobs)+1}',
                FormulaRule(formula=[f'F2<TODAY()'], stopIfTrue=True, fill=red_fill, font=red_font)
            )
            # Green if applied
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            green_font = Font(color='006100')
            ws_data.conditional_formatting.add(
                f'H2:H{len(jobs)+1}',
                CellIsRule(operator='equal', formula=['"Yes"'], stopIfTrue=True, fill=green_fill, font=green_font)
            )

            # Dashboard Sheet (Charts)
            ws_dash = wb.create_sheet("Analytics Dashboard", 0)
            ws_dash.sheet_properties.tabColor = "107C10"
            
            ws_dash['A1'] = "Status"
            ws_dash['B1'] = "Count"
            status_counts = {}
            for j, _ in jobs:
                status_counts[j.status] = status_counts.get(j.status, 0) + 1
            
            row = 2
            for s, c in status_counts.items():
                ws_dash.cell(row=row, column=1, value=s)
                ws_dash.cell(row=row, column=2, value=c)
                row += 1
                
            if status_counts:
                chart = PieChart()
                data = Reference(ws_dash, min_col=2, min_row=1, max_row=row-1)
                cats = Reference(ws_dash, min_col=1, min_row=2, max_row=row-1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.title = "Jobs by Status"
                ws_dash.add_chart(chart, "D2")
                
            wb.save(self.output_path)
            
            # Create Backup
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(self.backup_dir, f"GovTrackAI_Tracker_{timestamp}.xlsx")
            shutil.copy2(self.output_path, backup_path)
            
            return self.output_path
        finally:
            db.close()
""",

    "excel/importer.py": """import os
from openpyxl import load_workbook
from db.connection import SessionLocal
from db.models import Job

class ExcelImporter:
    def import_data(self, file_path):
        db = SessionLocal()
        try:
            wb = load_workbook(file_path, data_only=True)
            if "Recruitment Tracker" not in wb.sheetnames:
                raise ValueError("Missing 'Recruitment Tracker' sheet.")
            ws = wb["Recruitment Tracker"]
            
            headers = [cell.value for cell in ws[1]]
            if "ID" not in headers:
                raise ValueError("Missing 'ID' column.")
            
            id_idx = headers.index("ID")
            status_idx = headers.index("Status") if "Status" in headers else -1
            applied_idx = headers.index("Applied") if "Applied" in headers else -1
            pri_idx = headers.index("Priority") if "Priority" in headers else -1
            
            updated_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                job_id = row[id_idx]
                if not job_id:
                    continue
                
                job = db.query(Job).filter(Job.id == job_id).first()
                if job:
                    dirty = False
                    if status_idx != -1 and row[status_idx] and job.status != row[status_idx]:
                        job.status = str(row[status_idx])
                        dirty = True
                    if applied_idx != -1 and row[applied_idx] is not None:
                        is_app = str(row[applied_idx]).lower().strip() == 'yes'
                        if job.is_applied != is_app:
                            job.is_applied = is_app
                            dirty = True
                    if pri_idx != -1 and row[pri_idx] is not None:
                        try:
                            pri = int(row[pri_idx])
                            if job.priority != pri:
                                job.priority = pri
                                dirty = True
                        except: pass
                    
                    if dirty:
                        updated_count += 1
            
            db.commit()
            return updated_count
        finally:
            db.close()
""",

    "api/routers/excel_sync.py": """from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from excel.generator import ExcelGenerator
from excel.importer import ExcelImporter
import os
import shutil
import logging

logger = logging.getLogger('app.excel_sync')
router = APIRouter(prefix="/excel", tags=["Excel"])

@router.get("/export")
def export_excel():
    generator = ExcelGenerator()
    path = generator.generate_dashboard()
    return FileResponse(path, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename="GovTrackAI_Tracker.xlsx")

@router.post("/import")
def import_excel(file: UploadFile = File(...)):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(400, "Only .xlsx files are supported")
        
    temp_path = f"temp_{file.filename}"
    try:
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        importer = ExcelImporter()
        updated = importer.import_data(temp_path)
        
        # After import, regenerate the auto-backup to sync SQLite state back to excel completely
        ExcelGenerator().generate_dashboard()
        
        return {"status": "success", "updated_count": updated}
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)

@router.post("/backup")
def force_backup():
    generator = ExcelGenerator()
    path = generator.generate_dashboard()
    return {"status": "success", "message": "Backup created successfully"}
""",

    "api/main.py": """from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from api.routers import jobs, analytics, profile, calendar, excel_sync

app = FastAPI(title="GovTrack AI API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(jobs.router, prefix="/api/v1")
app.include_router(analytics.router, prefix="/api/v1")
app.include_router(profile.router, prefix="/api/v1")
app.include_router(calendar.router, prefix="/api/v1")
app.include_router(excel_sync.router, prefix="/api/v1")

app.mount("/", StaticFiles(directory="frontend", html=True), name="frontend")
"""
}

for filepath, content in files.items():
    full_path = os.path.join(project_root, filepath)
    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    with open(full_path, 'w') as f:
        f.write(content)

print("Phase 25 Backend Excel complete.")
