import os
from openpyxl import load_workbook
from db.connection import SessionLocal
from db.models import Job

class ExcelImporter:
    def import_data(self, file_path):
        db = SessionLocal()
        try:
            wb = load_workbook(file_path, data_only=True)
            if "Recruitment Tracker" not in wb.sheetnames:
                raise ValueError("Missing 'Recruitment Tracker' sheet.")
            ws = wb["Recruitment Tracker"]
            
            headers = [cell.value for cell in ws[1]]
            if "ID" not in headers:
                raise ValueError("Missing 'ID' column.")
            
            id_idx = headers.index("ID")
            status_idx = headers.index("Status") if "Status" in headers else -1
            applied_idx = headers.index("Applied") if "Applied" in headers else -1
            pri_idx = headers.index("Priority") if "Priority" in headers else -1
            
            updated_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                job_id = row[id_idx]
                if not job_id:
                    continue
                
                job = db.query(Job).filter(Job.id == job_id).first()
                if job:
                    dirty = False
                    if status_idx != -1 and row[status_idx] and job.status != row[status_idx]:
                        job.status = str(row[status_idx])
                        dirty = True
                    if applied_idx != -1 and row[applied_idx] is not None:
                        is_app = str(row[applied_idx]).lower().strip() == 'yes'
                        if job.is_applied != is_app:
                            job.is_applied = is_app
                            dirty = True
                    if pri_idx != -1 and row[pri_idx] is not None:
                        try:
                            pri = int(row[pri_idx])
                            if job.priority != pri:
                                job.priority = pri
                                dirty = True
                        except: pass
                    
                    if dirty:
                        updated_count += 1
            
            db.commit()
            return updated_count
        finally:
            db.close()
